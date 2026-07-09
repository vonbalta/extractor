"""
report_builder.py — Construcción y escritura de la hoja 'Desviaciones'.

Responsabilidades
-----------------
- Recorrer las hojas de un workbook openpyxl y recolectar las desviaciones
  detectadas por deviation_rules.
- Poblar una hoja openpyxl con encabezados, datos, formatos y fusiones de
  celdas según el layout de la hoja 'Desviaciones'.
- Preservar los datos rellenados manualmente (cols F, G, H) cuando la hoja
  se regenera.
- Crear o reemplazar la hoja 'Desviaciones' en un workbook openpyxl
  (función de compatibilidad `escribir_hoja`).

No tiene dependencias de xlwings, PDF ni de interfaz gráfica.
"""

import logging
from datetime import datetime, date

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pdf_extractor import MARCAS_AL_FINAL
from excel_manager import PATRON_HOJA_NUM
import deviation_rules as dr

logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())


# ---------------------------------------------------------------------------
#  HELPERS DE FORMATO (módulo-privados)
# ---------------------------------------------------------------------------

def _fmt_fecha(val):
    """Convierte datetime/date a 'DD/MM/YYYY' para escritura segura en celda."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    return str(val)


def _clave_normalizada(fecha_val, ciclo_val, auto_val):
    """
    Normaliza la terna (fecha, ciclo, autoclave) a una tupla de strings estable,
    usada como clave para emparejar filas entre la hoja 'Desviaciones' anterior
    y la recién generada.

    Centraliza la lógica que antes estaba duplicada en `leer_contenido_existente`
    y `aplicar_contenido_existente`; cualquier cambio de formato de clave se hace
    ahora en un solo lugar, evitando que ambas copias se desincronicen (lo que
    rompería silenciosamente la preservación de datos manuales).
    """
    # — Fecha —
    if isinstance(fecha_val, datetime):
        fecha_k = fecha_val.strftime("%d/%m/%Y")
    elif isinstance(fecha_val, date):
        fecha_k = fecha_val.strftime("%d/%m/%Y")
    elif isinstance(fecha_val, str):
        fecha_k = fecha_val.strip()
    else:
        fecha_k = str(fecha_val).strip() if fecha_val is not None else ""

    # — Ciclo (un float entero '3.0' debe casar con el int '3') —
    if isinstance(ciclo_val, bool):
        # bool es subclase de int; se trata como texto para no producir '0'/'1'.
        ciclo_k = str(ciclo_val)
    elif isinstance(ciclo_val, float):
        if ciclo_val == ciclo_val and ciclo_val == int(ciclo_val):  # descarta NaN
            ciclo_k = str(int(ciclo_val))
        else:
            ciclo_k = str(ciclo_val).strip()
    elif isinstance(ciclo_val, int):
        ciclo_k = str(ciclo_val)
    else:
        ciclo_k = str(ciclo_val).strip() if ciclo_val is not None else ""

    # — Autoclave —
    auto_k = str(auto_val).strip() if auto_val is not None else ""

    return (fecha_k, ciclo_k, auto_k)


# ---------------------------------------------------------------------------
#  RECOLECCIÓN DE DESVIACIONES
# ---------------------------------------------------------------------------

def recolectar_desviaciones(wb_opx, evento_cancelacion=None,
                             verificar_cancelacion_fn=None):
    """
    Recorre todas las hojas del workbook openpyxl (excepto 'Desviaciones') y
    devuelve una lista de dicts, uno por hoja con desviaciones detectadas.

    Cada dict contiene:
        fecha, desc, hora_desv, ciclo, ciclo_texto, autoclave_texto, es_marca

    El formato del texto (caso único vs. repetido) se decide aquí:
      - Una sola ocurrencia → oración completa con lecturas + hora en 'hora_desv'.
      - Varias ocurrencias  → encabezado + una línea por ocurrencia; 'hora_desv' vacío.

    Parámetros opcionales para cancelación cooperativa:
        evento_cancelacion       : threading.Event
        verificar_cancelacion_fn : callable(evento) → lanza OperacionCancelada
    """
    registros     = []
    contador_marca = 0
    hojas         = [n for n in wb_opx.sheetnames if n != "Desviaciones"]
    logger.info(f"Hojas a analizar: {len(hojas)}")

    for ws_name in hojas:
        if verificar_cancelacion_fn:
            verificar_cancelacion_fn(evento_cancelacion)
        ws       = wb_opx[ws_name]
        df_filas = dr.leer_filas_hoja(ws)
        logger.info(f"  · {ws_name}: {len(df_filas)} fila(s) con datos")

        fecha_ciclo = dr.fecha_hoja(ws)
        entradas    = dr.evaluar_hoja(df_filas)

        if not entradas:
            continue

        # — Formatear descripciones y horas de desviación —
        descripciones = []
        horas_desv    = []

        for ent in entradas:
            ocs = ent["ocurrencias"]
            if len(ocs) == 1:
                # Ocurrencia única: oración con lecturas en línea; hora aparte.
                oc = ocs[0]
                descripciones.append(oc.get("oracion", ent["header"]))
                horas_desv.append(oc["hora_corta"])
            else:
                # Ocurrencia repetida: encabezado + una línea por hora/lectura.
                if ent["con_lecturas"]:
                    lineas = [f"{oc['hora_corta']} {oc['lecturas']}" for oc in ocs]
                else:
                    lineas = [oc["hora_corta"] for oc in ocs]
                descripciones.append(ent["header"] + "\n" + "\n".join(lineas))
                horas_desv.append("")

        # — Extraer metadatos del nombre de hoja —
        match_num     = PATRON_HOJA_NUM.match(ws_name)
        num_ciclo     = int(match_num.group(1)) if match_num else None
        num_autoclave = int(match_num.group(2)) if match_num else None
        autoclave_txt = num_autoclave if num_autoclave is not None else ""

        marca_nombre = (
            match_num.group(3).lower()
            if (match_num and match_num.group(3))
            else None
        )
        es_marca = marca_nombre in MARCAS_AL_FINAL
        if es_marca:
            contador_marca += 1
            ciclo_texto = f"{contador_marca} Maquila"
        else:
            ciclo_texto = str(num_ciclo) if num_ciclo is not None else ""

        registros.append({
            "fecha":           fecha_ciclo,
            "desc":            "\n".join(descripciones),
            "hora_desv":       "\n".join(horas_desv),
            "ciclo":           num_ciclo,      # número real (orden/merge)
            "ciclo_texto":     ciclo_texto,    # texto a mostrar
            "autoclave_texto": autoclave_txt,
            "es_marca":        es_marca,
        })

    return registros


# ---------------------------------------------------------------------------
#  ESCRITURA DE LA HOJA (openpyxl)
# ---------------------------------------------------------------------------

def poblar_hoja(ws, registros):
    """
    Escribe encabezados, datos, formatos y fusiones de celdas en el
    openpyxl Worksheet `ws`.

    Se desacopla de la creación del workbook para poder usarse tanto en
    `escribir_hoja` (compatibilidad) como directamente sobre un workbook
    temporal de una sola hoja que luego se copia vía xlwings.
    """
    # — Estilos —
    font_header = Font(name="Arial", bold=True,  size=10)
    font_cell   = Font(name="Arial", bold=False, size=10)
    fill_header = PatternFill("solid", fgColor="D9E1F2")
    alin_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alin_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin")
    borde       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["Fecha", "Ciclo", "Autoclave", "Desviación", "Hora de desviación",
                  "Acción preventiva", "Cantidad de tinas retenidas", "Completado"]
    col_widths = [14, 8, 12, 62, 22, 22, 28, 14]

    # — Fila de encabezados —
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(row=1, column=ci, value=h)
        cell.font      = font_header
        cell.fill      = fill_header
        cell.alignment = alin_center
        cell.border    = borde
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30

    if not registros:
        return

    # Asegurar que cada registro tenga 'ciclo_texto' y 'autoclave_texto'.
    filas_datos = [
        {
            **rec,
            "ciclo":          rec["ciclo"] if rec.get("ciclo") is not None else i,
            "ciclo_texto":    rec.get("ciclo_texto") or str(rec.get("ciclo", i)),
            "autoclave_texto": rec.get("autoclave_texto", ""),
        }
        for i, rec in enumerate(registros, start=1)
    ]

    # — Filas de datos —
    for er, rec in enumerate(filas_datos, start=2):
        def _celda(col, valor, alin=alin_center, _er=er):
            cell           = ws.cell(row=_er, column=col, value=valor)
            cell.font      = font_cell
            cell.alignment = alin
            cell.border    = borde
            return cell

        _celda(1, _fmt_fecha(rec["fecha"]))
        _celda(2, rec["ciclo_texto"] if rec.get("es_marca") else rec["ciclo"])
        _celda(3, rec["autoclave_texto"])
        _celda(4, rec["desc"],     alin=alin_left)
        celda_e = _celda(5, rec["hora_desv"])
        celda_e.number_format = "@"   # fuerza texto; evita reinterpretación de "HH:MM"

        for cv in (6, 7, 8):          # Acción preventiva / tinas / completado
            ws.cell(row=er, column=cv).border = borde

        # Alto de fila proporcional al número de líneas del contenido.
        num_lineas = max(
            rec["desc"].count("\n") + 1,
            rec["hora_desv"].count("\n") + 1,
        )
        ws.row_dimensions[er].height = max(18, num_lineas * 15)

    # — Fusión de celdas por ciclo —
    # Columnas A, B, C, E, F, G, H se fusionan cuando un mismo ciclo ocupa
    # más de una fila (D se deja sin fusionar: contiene texto multilínea).
    columnas_a_fusionar = (1, 2, 3, 5, 6, 7, 8)

    i = 0
    while i < len(filas_datos):
        ciclo_actual = filas_datos[i]["ciclo"]
        inicio_excel = 2 + i
        j = i
        while j < len(filas_datos) and filas_datos[j]["ciclo"] == ciclo_actual:
            j += 1
        fin_excel = 2 + j - 1

        if fin_excel > inicio_excel:
            for col in columnas_a_fusionar:
                ws.merge_cells(
                    start_row=inicio_excel, start_column=col,
                    end_row=fin_excel,      end_column=col,
                )

        # Estilo en la celda superior del grupo (merged o no).
        for col in columnas_a_fusionar:
            mc           = ws.cell(row=inicio_excel, column=col)
            mc.font      = font_cell
            mc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            mc.border    = borde
            if col == 5:
                mc.number_format = "@"

        i = j


# ---------------------------------------------------------------------------
#  PRESERVAR DATOS MANUALES
# ---------------------------------------------------------------------------

def leer_contenido_existente(wb_opx):
    """
    Lee la hoja 'Desviaciones' existente y devuelve un dict indexado por
    (fecha_str, ciclo_str, autoclave_str) con los valores de las columnas F,
    G y H (Acción preventiva, Tinas, Completado) rellenados manualmente.

    Solo se conservan filas donde al menos una de las tres columnas tiene
    contenido distinto de vacío/None.
    """
    contenido = {}
    if "Desviaciones" not in wb_opx.sheetnames:
        return contenido

    ws = wb_opx["Desviaciones"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue

        fecha_val, ciclo_val, auto_val = row[0], row[1], row[2]
        accion, tinas, completado      = row[5], row[6], row[7]

        if not any(
            v is not None and str(v).strip() != ""
            for v in (accion, tinas, completado)
        ):
            continue

        key = _clave_normalizada(fecha_val, ciclo_val, auto_val)
        if key not in contenido:
            contenido[key] = {"accion": accion, "tinas": tinas, "completado": completado}

    return contenido


def aplicar_contenido_existente(ws, contenido_previo):
    """
    Rellena las columnas F, G y H de la hoja 'Desviaciones' recién generada
    con los valores guardados manualmente en la versión anterior, emparejando
    por la clave (Fecha, Ciclo, Autoclave).
    """
    if not contenido_previo:
        return

    for row_idx in range(2, ws.max_row + 1):
        fecha_val = ws.cell(row=row_idx, column=1).value
        ciclo_val = ws.cell(row=row_idx, column=2).value
        auto_val  = ws.cell(row=row_idx, column=3).value

        if fecha_val is None and ciclo_val is None and auto_val is None:
            continue

        key = _clave_normalizada(fecha_val, ciclo_val, auto_val)
        if key in contenido_previo:
            datos = contenido_previo[key]
            if datos.get("accion") is not None:
                ws.cell(row=row_idx, column=6).value = datos["accion"]
            if datos.get("tinas") is not None:
                ws.cell(row=row_idx, column=7).value = datos["tinas"]
            if datos.get("completado") is not None:
                ws.cell(row=row_idx, column=8).value = datos["completado"]


# ---------------------------------------------------------------------------
#  COMPATIBILIDAD: escribir hoja directamente en un workbook openpyxl
# ---------------------------------------------------------------------------

def escribir_hoja(wb_opx, registros):
    """
    Crea (o reemplaza) la hoja 'Desviaciones' en `wb_opx` y delega el
    relleno en `poblar_hoja`.

    NOTA: usar esta función directamente implica que openpyxl guarda el
    workbook completo, lo que puede eliminar drawings y vínculos externos.
    Para el flujo principal (Modo 3) se prefiere el approach de xlwings
    descrito en `logica.ejecutar_revisor_desviaciones`.
    """
    if "Desviaciones" in wb_opx.sheetnames:
        del wb_opx["Desviaciones"]
    ws = wb_opx.create_sheet("Desviaciones")
    poblar_hoja(ws, registros)